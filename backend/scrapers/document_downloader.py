import asyncio
import hashlib
import os
import io
import re
from datetime import datetime
from pathlib import Path

import httpx
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.database import SessionLocal, get_async_db
from app.models import TenderDocument, Tender, DocumentDownload
from app.config import settings

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    import docx
except ImportError:
    docx = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

DOWNLOADS_BASE_DIR = Path("downloads")
TEXT_LIMIT = 1_000_000


def _index_tender_sync(tender_id: int) -> None:
    from app.services.tender_index import index_tender

    db = SessionLocal()
    try:
        tender = db.get(Tender, tender_id)
        if tender:
            index_tender(db, tender, commit=True)
    finally:
        db.close()

class DocumentDownloader:
    def __init__(self, concurrency: int = 5):
        self.concurrency = concurrency
        self.semaphore = asyncio.Semaphore(concurrency)

    async def run(self):
        DOWNLOADS_BASE_DIR.mkdir(parents=True, exist_ok=True)
        
        async for db in get_async_db():
            result = await db.execute(
                select(TenderDocument.id)
                .where(TenderDocument.status.in_(["queued", "retrying"]))
            )
            download_ids = result.scalars().all()
            break

        tasks = [asyncio.create_task(self._process_download(doc_id)) for doc_id in download_ids]
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

    async def _process_download(self, doc_id: int):
        async with self.semaphore:
            async for db in get_async_db():
                await self._process_download_with_session(db, doc_id)
                break

    async def _process_download_with_session(self, db: AsyncSession, doc_id: int):
            result = await db.execute(select(TenderDocument).where(TenderDocument.id == doc_id))
            doc = result.scalars().first()
            if not doc or doc.status not in {"queued", "retrying"}:
                return
            doc.status = "processing"
            await db.commit()
            
            res = await db.execute(select(Tender).where(Tender.id == doc.tender_id))
            tender = res.scalars().first()
            if not tender:
                doc.status = "failed"
                doc.error_message = "Tender not found in database"
                await db.commit()
                return

            portal_clean = re.sub(r'[^a-zA-Z0-9_-]', '_', tender.portal or "unknown")
            tender_id_clean = re.sub(r'[^a-zA-Z0-9_-]', '_', tender.tender_id or "unknown")
            
            portal_dir = DOWNLOADS_BASE_DIR / portal_clean / tender_id_clean
            portal_dir.mkdir(parents=True, exist_ok=True)
            
            url_path = doc.url.split('?')[0]
            filename = os.path.basename(url_path) or f"doc_{doc.id}"
            filename = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
            file_path = portal_dir / filename
            
            try:
                # Retries
                content = b""
                last_err = None
                for attempt in range(3):
                    try:
                        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True, verify=False) as client:
                            response = await client.get(doc.url)
                            response.raise_for_status()
                            content = response.content
                            break
                    except Exception as e:
                        last_err = e
                        await asyncio.sleep(2.0 * (attempt + 1))
                
                if not content:
                    raise last_err or RuntimeError("Failed to download document content")
                
                file_hash = hashlib.sha256(content).hexdigest()
                file_path.write_bytes(content)
                
                doc.file_name = filename
                doc.storage_path = str(file_path.resolve())
                doc.content_hash = file_hash
                doc.processed_at = datetime.utcnow()
                
                # Extract text
                file_ext = os.path.splitext(filename)[1].lower().lstrip('.')
                extracted_text = self._extract_text(content, file_ext, file_path)
                doc.extracted_text = extracted_text
                doc.status = "processed" if extracted_text else "downloaded"
                doc.error_message = None
                
                if extracted_text:
                    tender.search_text = (tender.search_text or "") + "\n" + extracted_text
                    raw_data = dict(tender.raw_data or {})
                    existing_doc_text = raw_data.get("document_text") or ""
                    raw_data["document_text"] = f"{existing_doc_text}\n{extracted_text}"[:TEXT_LIMIT] if existing_doc_text else extracted_text[:TEXT_LIMIT]
                    tender.raw_data = raw_data
                    tender.classification_status = "PENDING_CLASSIFICATION"
                
                # Update/Create DocumentDownload
                dd_res = await db.execute(
                    select(DocumentDownload)
                    .where(DocumentDownload.tender_id == tender.id, DocumentDownload.url == doc.url)
                )
                dd = dd_res.scalars().first()
                if not dd:
                    dd = DocumentDownload(
                        tender_id=tender.id,
                        url=doc.url,
                    )
                    db.add(dd)
                
                dd.filename = filename
                dd.file_type = file_ext
                dd.file_size = len(content)
                dd.checksum = file_hash
                dd.storage_path = str(file_path.resolve())
                dd.status = "completed"
                dd.downloaded_at = datetime.utcnow()
                dd.error_message = None
                
                await db.commit()
                if extracted_text:
                    await asyncio.to_thread(_index_tender_sync, tender.id)
                
            except Exception as e:
                doc.status = "failed"
                doc.error_message = str(e)
                
                dd_res = await db.execute(
                    select(DocumentDownload)
                    .where(DocumentDownload.tender_id == tender.id, DocumentDownload.url == doc.url)
                )
                dd = dd_res.scalars().first()
                if not dd:
                    dd = DocumentDownload(
                        tender_id=tender.id,
                        url=doc.url,
                    )
                    db.add(dd)
                dd.status = "failed"
                dd.error_message = str(e)
                dd.downloaded_at = datetime.utcnow()
                
                await db.commit()


    def _extract_text(self, content: bytes, file_type: str, path: Path) -> str:
        text = ""
        try:
            if file_type == "pdf" and PdfReader:
                reader = PdfReader(io.BytesIO(content))
                text = "\n".join(page.extract_text() or "" for page in reader.pages)[:TEXT_LIMIT]
            elif file_type in {"xlsx", "xlsm"} and load_workbook:
                workbook = load_workbook(path, read_only=True, data_only=True)
                lines = []
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        values = [str(value) for value in row if value not in (None, "")]
                        if values:
                            lines.append(" ".join(values))
                        if sum(len(line) for line in lines) > TEXT_LIMIT:
                            break
                text = "\n".join(lines)[:TEXT_LIMIT]
            elif file_type == "docx" and docx:
                doc_obj = docx.Document(io.BytesIO(content))
                text = "\n".join(p.text for p in doc_obj.paragraphs)[:TEXT_LIMIT]
            elif file_type in {"txt", "csv"}:
                text = content.decode("utf-8", errors="ignore")[:TEXT_LIMIT]
        except Exception as e:
            print(f"Failed to extract text from {file_type}: {e}")
            
        return text

DOCUMENT_DOWNLOADER = DocumentDownloader()
