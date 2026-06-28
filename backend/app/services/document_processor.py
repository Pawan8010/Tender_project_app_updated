import hashlib
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.models import Tender, TenderDocument


DOCUMENT_DIR = Path("data") / "documents"
TEXT_LIMIT = 1_000_000


def _safe_name(url: str, fallback: str) -> str:
    path_name = Path(urlparse(url).path).name or fallback
    cleaned = "".join(char if char.isalnum() or char in "._-" else "_" for char in path_name)
    return cleaned[:180] or fallback


def _hash_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""
    try:
        reader = PdfReader(str(path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)[:TEXT_LIMIT]
    except Exception:
        return ""


def _extract_xlsx(path: Path) -> str:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value not in (None, "")]
                if values:
                    lines.append(" ".join(values))
                if sum(len(line) for line in lines) > TEXT_LIMIT:
                    break
        return "\n".join(lines)[:TEXT_LIMIT]
    except Exception:
        return ""


def _extract_text_file(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")[:TEXT_LIMIT]
    except Exception:
        return ""


def _extract_zip_listing(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            return "\n".join(archive.namelist())[:TEXT_LIMIT]
    except Exception:
        return ""


def extract_document_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_xlsx(path)
    if suffix in {".txt", ".csv"}:
        return _extract_text_file(path)
    if suffix == ".zip":
        return _extract_zip_listing(path)
    return ""


def _merge_tender_search_text(tender: Tender, text: str) -> None:
    if not text:
        return
    existing = tender.search_text or ""
    merged = f"{existing}\n{text}" if existing else text
    tender.search_text = merged[:TEXT_LIMIT]
    raw_data = dict(tender.raw_data or {})
    existing_doc_text = raw_data.get("document_text") or ""
    raw_data["document_text"] = f"{existing_doc_text}\n{text}"[:TEXT_LIMIT] if existing_doc_text else text[:TEXT_LIMIT]
    tender.raw_data = raw_data
    tender.classification_status = "PENDING_CLASSIFICATION"
    tender.updated_at = datetime.utcnow()


def process_queued_documents(db: Session, limit: int = 20) -> dict:
    DOCUMENT_DIR.mkdir(parents=True, exist_ok=True)
    docs = (
        db.query(TenderDocument)
        .filter(TenderDocument.status.in_(["queued", "retrying"]))
        .order_by(TenderDocument.created_at.asc())
        .limit(limit)
        .all()
    )
    processed = 0
    failed = 0
    skipped = 0
    for doc in docs:
        doc.status = "processing"
        db.flush()
        try:
            with httpx.Client(timeout=settings()["scraper_request_timeout_seconds"], follow_redirects=True, verify=False) as client:
                response = client.get(doc.url, headers={"User-Agent": "Mozilla/5.0 TenderIntel/1.0"})
                response.raise_for_status()
                content = response.content
            digest = _hash_bytes(content)
            file_name = doc.file_name or _safe_name(doc.url, f"document-{doc.id}")
            storage_name = f"{doc.id}-{digest[:12]}-{_safe_name(doc.url, file_name)}"
            storage_path = DOCUMENT_DIR / storage_name
            storage_path.write_bytes(content)
            text = extract_document_text(storage_path)
            doc.storage_path = str(storage_path)
            doc.content_hash = digest
            doc.extracted_text = text
            doc.status = "processed" if text or content else "downloaded"
            doc.processed_at = datetime.utcnow()
            doc.error_message = None
            if doc.tender:
                _merge_tender_search_text(doc.tender, text)
            processed += 1
        except Exception as exc:
            doc.status = "failed"
            doc.error_message = str(exc)[:1000]
            doc.processed_at = datetime.utcnow()
            failed += 1
        db.commit()
    return {"processed": processed, "failed": failed, "skipped": skipped, "remaining": max(0, len(docs) - processed - failed - skipped)}
